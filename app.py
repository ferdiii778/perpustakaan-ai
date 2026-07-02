from flask import Flask, render_template, request, redirect, flash, session
import psycopg2
from werkzeug.utils import secure_filename
import os
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from supabase import create_client
import psycopg2.extras


app = Flask(__name__)
app.secret_key = "perpustakaan123"
SUPABASE_URL = "https://gucdxbumigjiqlpbzxbd.supabase.co"
SUPABASE_SERVICE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Imd1Y2R4YnVtaWdqaXFscGJ6eGJkIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4MTQzOTc2OSwiZXhwIjoyMDk3MDE1NzY5fQ.90WoraIjPbbslKzR8ZWny_ULtlaeoK_awL4soVgAa_E"

supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

PROFILE_UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'images')
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'images', 'books')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(PROFILE_UPLOAD_FOLDER, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Koneksi Supabase PostgreSQL
def get_db_connection():
    return psycopg2.connect(
        host="aws-1-ap-northeast-1.pooler.supabase.com",
        database="postgres",
        user="postgres.gucdxbumigjiqlpbzxbd",
        password="Al_iryadIu*",
        port="5432",
        sslmode="require"
    )

@app.route('/')
def login():
    return render_template('auth/login.html')

def upload_sampul_ke_supabase(file):
    if not file or file.filename == "":
        return ""

    nama_file = secure_filename(file.filename)
    file_path_storage = f"books/{nama_file}"
    file_bytes = file.read()

    supabase.storage.from_("book-covers").upload(
        file_path_storage,
        file_bytes,
        {
            "content-type": file.content_type,
            "upsert": "true"
        }
    )

    return supabase.storage.from_("book-covers").get_public_url(file_path_storage)

@app.route('/proses_login', methods=['POST'])
def proses_login():
    email = request.form['email']
    password = request.form['password']

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id_user, nama, email, password, role
        FROM users
        WHERE email=%s AND password=%s AND aktif=true
    """, (email, password))

    user = cursor.fetchone()

    cursor.close()
    conn.close()

    if user:
        session['login'] = True
        session['id_user'] = user[0]
        session['nama'] = user[1]
        session['email'] = user[2]
        session['role'] = user[4]

        flash('Login Berhasil!', 'success')

        if session['role'] == 'admin':
            return redirect('/dashboard_admin')
        else:
            return redirect('/dashboard_user')

    flash('Email atau password salah!', 'danger')
    return redirect('/')

@app.route('/dashboard_admin')
def dashboard_admin():
    if 'login' not in session:
        return redirect('/')

    if session['role'] != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect('/dashboard_user')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM buku")
    total_buku = cursor.fetchone()[0]

    cursor.execute("SELECT COALESCE(SUM(stok_tersedia), 0) FROM buku")
    tersedia = cursor.fetchone()[0]

    cursor.execute("""
        SELECT *
        FROM buku
        ORDER BY id_buku DESC
        LIMIT 8
    """)
    buku = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        'admin/dashboard_admin.html',
        nama=session['nama'],
        total_buku=total_buku,
        tersedia=tersedia,
        buku=buku
    )

@app.route('/dashboard_user')
def dashboard_user():
    if 'login' not in session:
        return redirect('/')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM buku")
    total_buku = cursor.fetchone()[0]

    cursor.execute("SELECT COALESCE(SUM(stok_tersedia), 0) FROM buku")
    tersedia = cursor.fetchone()[0]

    cursor.execute("""
        SELECT *
        FROM buku
        ORDER BY id_buku DESC
        LIMIT 10
    """)
    buku = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        'member/dashboard_user.html',
        nama=session['nama'],
        buku=buku,
        total_buku=total_buku,
        tersedia=tersedia
    )

@app.route('/dashboard')
def dashboard():
    if 'login' not in session:
        return redirect('/')

    if session['role'] == 'admin':
        return redirect('/dashboard_admin')

    return redirect('/dashboard_user')


@app.route('/kelola_buku')
def kelola_buku():
    if 'login' not in session:
        return redirect('/')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM buku")
    total_buku = cursor.fetchone()[0]

    cursor.execute("SELECT COALESCE(SUM(stok_tersedia), 0) FROM buku")
    tersedia = cursor.fetchone()[0]

    cursor.execute("SELECT COALESCE(SUM(jumlah_eksemplar - stok_tersedia), 0) FROM buku")
    dipinjam = cursor.fetchone()[0]

    cursor.execute("""
        SELECT *
        FROM buku
        ORDER BY id_buku DESC
    """)

    buku = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        'admin/kelola_buku.html',
        nama=session['nama'],
        buku=buku,
        total_buku=total_buku,
        tersedia=tersedia,
        dipinjam=dipinjam
    )


@app.route('/tambah_buku')
def tambah_buku():
    if 'login' not in session:
        return redirect('/')

    return render_template(
        'admin/tambah_buku.html',
        nama=session['nama'],
        mode='tambah',
        buku=None
    )

@app.route('/edit_buku/<int:id>')
def edit_buku(id):
    if 'login' not in session:
        return redirect('/')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM buku
        WHERE id_buku = %s
    """, (id,))

    buku = cursor.fetchone()

    cursor.close()
    conn.close()

    if buku is None:
        flash('Data buku tidak ditemukan', 'danger')
        return redirect('/kelola_buku')

    return render_template(
        'admin/tambah_buku.html',
        nama=session['nama'],
        mode='edit',
        buku=buku
    )

@app.route('/daftar_buku')
def daftar_buku():
    if 'login' not in session:
        return redirect('/')

    q = request.args.get('q', '').strip()

    page = request.args.get('page', 1, type=int)
    per_page = 12
    offset = (page - 1) * per_page

    conn = get_db_connection()
    cursor = conn.cursor()

    if q:
        keyword = f"%{q}%"

        cursor.execute("""
            SELECT COUNT(*)
            FROM buku
            WHERE
                judul_buku ILIKE %s OR
                penulis ILIKE %s OR
                penerbit ILIKE %s OR
                isbn ILIKE %s OR
                bahasa ILIKE %s OR
                LOWER(kategori) = LOWER(%s) OR
                LOWER(genre) = LOWER(%s) OR
                deskripsi ILIKE %s OR
                tags ILIKE %s
        """, (
            keyword, keyword, keyword,
            keyword, keyword, q,
            q, keyword, keyword
        ))

        total_buku = cursor.fetchone()[0]

        cursor.execute("""
            SELECT *
            FROM buku
            WHERE
                judul_buku ILIKE %s OR
                penulis ILIKE %s OR
                penerbit ILIKE %s OR
                isbn ILIKE %s OR
                bahasa ILIKE %s OR
                LOWER(kategori) = LOWER(%s) OR
                LOWER(genre) = LOWER(%s) OR
                deskripsi ILIKE %s OR
                tags ILIKE %s
            ORDER BY id_buku DESC
            LIMIT %s OFFSET %s
        """, (
            keyword, keyword, keyword,
            keyword, keyword, q,
            q, keyword, keyword,
            per_page, offset
        ))

    else:
        cursor.execute("SELECT COUNT(*) FROM buku")
        total_buku = cursor.fetchone()[0]

        cursor.execute("""
            SELECT *
            FROM buku
            ORDER BY id_buku DESC
            LIMIT %s OFFSET %s
        """, (per_page, offset))

    buku = cursor.fetchall()

    cursor.close()
    conn.close()

    total_pages = (total_buku + per_page - 1) // per_page

    return render_template(
        'member/daftar_buku.html',
        nama=session['nama'],
        buku=buku,
        total_buku=total_buku,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        q=q
    )

@app.route('/update_buku/<int:id>', methods=['POST'])
def update_buku(id):
    if 'login' not in session:
        return redirect('/')

    judul = request.form['judul_buku']
    penulis = request.form['penulis']
    penerbit = request.form['penerbit']
    tahun = request.form['tahun_terbit']
    isbn = request.form['isbn']
    bahasa = request.form['bahasa']
    kategori = request.form['kategori']
    genre = request.form['genre']
    deskripsi = request.form['deskripsi']
    jumlah = request.form['jumlah_eksemplar']
    stok = request.form['stok_tersedia']
    status = request.form['status_buku']
    rak = request.form['rak_lokasi']
    tags = request.form['tags']

    file = request.files.get('sampul')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT sampul
        FROM buku
        WHERE id_buku = %s
    """, (id,))

    old_data = cursor.fetchone()

    if old_data is None:
        cursor.close()
        conn.close()
        flash('Data buku tidak ditemukan', 'danger')
        return redirect('/kelola_buku')

    nama_file = old_data[0]

    if file and file.filename != "":
        nama_file = upload_sampul_ke_supabase(file)

    cursor.execute("""
        UPDATE buku
        SET
            judul_buku = %s,
            penulis = %s,
            penerbit = %s,
            tahun_terbit = %s,
            isbn = %s,
            bahasa = %s,
            kategori = %s,
            genre = %s,
            deskripsi = %s,
            sampul = %s,
            jumlah_eksemplar = %s,
            stok_tersedia = %s,
            status_buku = %s,
            rak_lokasi = %s,
            tags = %s
        WHERE id_buku = %s
    """, (
        judul, penulis, penerbit, tahun, isbn,
        bahasa, kategori, genre, deskripsi, nama_file,
        jumlah, stok, status, rak, tags, id
    ))

    conn.commit()
    cursor.close()
    conn.close()

    flash('Buku berhasil diperbarui', 'success')
    return redirect('/kelola_buku')

@app.route('/simpan_buku', methods=['POST'])
def simpan_buku():
    if 'login' not in session:
        return redirect('/')

    judul = request.form['judul_buku']
    penulis = request.form['penulis']
    penerbit = request.form['penerbit']
    tahun = request.form['tahun_terbit']
    isbn = request.form['isbn']
    bahasa = request.form['bahasa']
    kategori = request.form['kategori']
    genre = request.form['genre']
    deskripsi = request.form['deskripsi']
    jumlah = request.form['jumlah_eksemplar']
    stok = request.form['stok_tersedia']
    status = request.form['status_buku']
    rak = request.form['rak_lokasi']
    tags = request.form['tags']

    file = request.files.get('sampul')
    nama_file = upload_sampul_ke_supabase(file)

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO buku(
            judul_buku, penulis, penerbit, tahun_terbit, isbn,
            bahasa, kategori, genre, deskripsi, sampul,
            jumlah_eksemplar, stok_tersedia, status_buku,
            rak_lokasi, tags
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        judul, penulis, penerbit, tahun, isbn,
        bahasa, kategori, genre, deskripsi, nama_file,
        jumlah, stok, status, rak, tags
    ))

    conn.commit()
    cursor.close()
    conn.close()

    flash('Buku berhasil ditambahkan', 'success')
    return redirect('/kelola_buku')


@app.route('/detail_buku/<int:id>')
def detail_buku(id):
    if 'login' not in session:
        return redirect('/')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM buku WHERE id_buku=%s", (id,))
    buku = cursor.fetchone()

    cursor.close()
    conn.close()

    if buku is None:
        flash('Data buku tidak ditemukan', 'danger')
        return redirect('/daftar_buku')

    return render_template(
        'member/detail_buku.html',
        buku=buku,
        nama=session['nama']
    )

@app.route('/hapus_buku/<int:id>')
def hapus_buku(id):
    if 'login' not in session:
        return redirect('/')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute(
        "DELETE FROM buku WHERE id_buku=%s",
        (id,)
    )

    conn.commit()
    cursor.close()
    conn.close()

    flash('Buku berhasil dihapus', 'success')
    return redirect('/kelola_buku')


@app.route('/register')
def register():
    return render_template('auth/register.html')


@app.route('/riwayat')
def riwayat():
    if 'login' not in session:
        return redirect('/')

    conn = get_db_connection()
    cursor = conn.cursor()

    id_user = session['id_user']

    cursor.execute("""
        SELECT 
            p.id_peminjaman,
            b.judul_buku,
            b.penulis,
            b.sampul,
            p.tanggal_pinjam,
            p.tanggal_jatuh_tempo,
            p.tanggal_kembali,
            CASE
                WHEN p.tanggal_kembali IS NULL 
                    AND CURRENT_DATE > p.tanggal_jatuh_tempo
                    THEN (CURRENT_DATE - p.tanggal_jatuh_tempo) * COALESCE(pp.denda_per_hari, 1000)
                ELSE COALESCE(p.denda, 0)
            END AS denda,
            CASE
                WHEN p.tanggal_kembali IS NULL 
                    AND CURRENT_DATE > p.tanggal_jatuh_tempo
                    THEN 'terlambat'
                ELSE p.status_peminjaman
            END AS status_peminjaman
        FROM peminjaman p
        JOIN buku b ON p.id_buku = b.id_buku
        JOIN users u ON p.id_user = u.id_user
        LEFT JOIN pengaturan_peminjaman pp ON pp.role = u.role::text
        WHERE p.id_user = %s
        ORDER BY p.tanggal_pinjam DESC
    """, (id_user,))
    riwayat = cursor.fetchall()

    cursor.execute("""
        SELECT COUNT(*)
        FROM peminjaman
        WHERE id_user = %s
    """, (id_user,))
    total_peminjaman = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM peminjaman
        WHERE id_user = %s 
        AND status_peminjaman IN ('dipinjam', 'terlambat')
    """, (id_user,))
    buku_dipinjam = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COALESCE(SUM(
            CASE
                WHEN p.tanggal_kembali IS NULL 
                    AND CURRENT_DATE > p.tanggal_jatuh_tempo
                    THEN (CURRENT_DATE - p.tanggal_jatuh_tempo) * COALESCE(pp.denda_per_hari, 1000)
                ELSE COALESCE(p.denda, 0)
            END
        ), 0)
        FROM peminjaman p
        JOIN users u ON p.id_user = u.id_user
        LEFT JOIN pengaturan_peminjaman pp ON pp.role = u.role::text
        WHERE p.id_user = %s
    """, (id_user,))
    total_denda = cursor.fetchone()[0]

    cursor.execute("""
        SELECT 
            b.judul_buku,
            b.penulis,
            b.sampul,
            p.tanggal_jatuh_tempo
        FROM peminjaman p
        JOIN buku b ON p.id_buku = b.id_buku
        WHERE p.id_user = %s
        AND p.status_peminjaman IN ('dipinjam', 'terlambat')
        ORDER BY p.tanggal_jatuh_tempo ASC
        LIMIT 1
    """, (id_user,))
    jatuh_tempo = cursor.fetchone()

    cursor.close()
    conn.close()

    return render_template(
        'member/riwayat.html',
        nama=session['nama'],
        riwayat=riwayat,
        total_peminjaman=total_peminjaman,
        buku_dipinjam=buku_dipinjam,
        total_denda=total_denda,
        jatuh_tempo=jatuh_tempo
    )

@app.route('/kelola_pengembalian')
def kelola_pengembalian():
    if 'login' not in session:
        return redirect('/')

    if session['role'] != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect('/dashboard_user')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT 
            p.id_peminjaman,
            u.nama,
            u.nomor_induk,
            b.judul_buku,
            b.penulis,
            b.sampul,
            p.tanggal_pinjam,
            p.tanggal_jatuh_tempo,

            CASE
                WHEN p.tanggal_kembali IS NULL
                     AND CURRENT_DATE > p.tanggal_jatuh_tempo
                    THEN (CURRENT_DATE - p.tanggal_jatuh_tempo) * COALESCE(pp.denda_per_hari, 1000)
                ELSE COALESCE(p.denda, 0)
            END AS denda,

            CASE
                WHEN p.tanggal_kembali IS NULL
                     AND CURRENT_DATE > p.tanggal_jatuh_tempo
                    THEN 'terlambat'
                ELSE p.status_peminjaman
            END AS status_peminjaman,

            p.tanggal_kembali,

            CASE
                WHEN p.tanggal_kembali IS NULL
                     AND CURRENT_DATE > p.tanggal_jatuh_tempo
                    THEN CURRENT_DATE - p.tanggal_jatuh_tempo
                WHEN p.tanggal_kembali IS NOT NULL
                     AND p.tanggal_kembali > p.tanggal_jatuh_tempo
                    THEN p.tanggal_kembali - p.tanggal_jatuh_tempo
                ELSE 0
            END AS hari_terlambat

        FROM peminjaman p
        JOIN users u ON p.id_user = u.id_user
        JOIN buku b ON p.id_buku = b.id_buku
        LEFT JOIN pengaturan_peminjaman pp ON pp.role = u.role::text
        ORDER BY p.id_peminjaman DESC
    """)
    pengembalian = cursor.fetchall()

    cursor.execute("""
        SELECT COUNT(*)
        FROM peminjaman
        WHERE tanggal_kembali = CURRENT_DATE
    """)
    hari_ini = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM peminjaman
        WHERE status_peminjaman = 'dikembalikan'
    """)
    sudah_dikembalikan = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM peminjaman
        WHERE tanggal_kembali IS NULL
        AND CURRENT_DATE > tanggal_jatuh_tempo
    """)
    terlambat = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COALESCE(SUM(
            CASE
                WHEN p.tanggal_kembali IS NULL
                     AND CURRENT_DATE > p.tanggal_jatuh_tempo
                    THEN (CURRENT_DATE - p.tanggal_jatuh_tempo) * COALESCE(pp.denda_per_hari, 1000)
                ELSE COALESCE(p.denda, 0)
            END
        ), 0)
        FROM peminjaman p
        JOIN users u ON p.id_user = u.id_user
        LEFT JOIN pengaturan_peminjaman pp ON pp.role = u.role::text
    """)
    total_denda = cursor.fetchone()[0]

    cursor.close()
    conn.close()

    return render_template(
        'admin/kelola_pengembalian.html',
        nama=session['nama'],
        pengembalian=pengembalian,
        hari_ini=hari_ini,
        sudah_dikembalikan=sudah_dikembalikan,
        terlambat=terlambat,
        total_denda=total_denda
    )

@app.route('/export_laporan_denda_excel')
def export_laporan_denda_excel():
    if 'login' not in session:
        return redirect('/')

    if session['role'] != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect('/dashboard_user')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            p.id_peminjaman,
            u.nama,
            u.nomor_induk,
            COALESCE(u.kelas, '-') AS kelas,
            b.judul_buku,
            p.tanggal_pinjam,
            p.tanggal_jatuh_tempo,
            p.tanggal_kembali,
            CASE
                WHEN p.tanggal_kembali IS NULL
                     AND CURRENT_DATE > p.tanggal_jatuh_tempo
                    THEN CURRENT_DATE - p.tanggal_jatuh_tempo
                WHEN p.tanggal_kembali IS NOT NULL
                     AND p.tanggal_kembali > p.tanggal_jatuh_tempo
                    THEN p.tanggal_kembali - p.tanggal_jatuh_tempo
                ELSE 0
            END AS hari_terlambat,
            CASE
                WHEN p.tanggal_kembali IS NULL
                     AND CURRENT_DATE > p.tanggal_jatuh_tempo
                    THEN (CURRENT_DATE - p.tanggal_jatuh_tempo) * COALESCE(pp.denda_per_hari, 1000)
                ELSE COALESCE(p.denda, 0)
            END AS denda,
            CASE
                WHEN p.tanggal_kembali IS NULL
                     AND CURRENT_DATE > p.tanggal_jatuh_tempo
                    THEN 'Belum Dibayar'
                WHEN COALESCE(p.denda, 0) > 0
                    THEN 'Belum Dibayar'
                ELSE 'Tidak Ada Denda'
            END AS status_denda
        FROM peminjaman p
        JOIN users u ON p.id_user = u.id_user
        JOIN buku b ON p.id_buku = b.id_buku
        LEFT JOIN pengaturan_peminjaman pp ON pp.role = u.role::text
        WHERE
            (
                p.tanggal_kembali IS NULL
                AND CURRENT_DATE > p.tanggal_jatuh_tempo
            )
            OR COALESCE(p.denda, 0) > 0
        ORDER BY p.id_peminjaman DESC
    """)

    data = cursor.fetchall()

    cursor.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Denda"

    ws.merge_cells('A1:K1')
    ws['A1'] = 'LAPORAN DENDA PERPUSTAKAAN'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = [
        'No',
        'ID Denda',
        'Nama Anggota',
        'Nomor Induk',
        'Kelas',
        'Judul Buku',
        'Tanggal Pinjam',
        'Jatuh Tempo',
        'Tanggal Kembali',
        'Hari Terlambat',
        'Denda',
        'Status Denda'
    ]

    ws.append([])
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="16A34A")
    header_font = Font(color="FFFFFF", bold=True)

    border = Border(
        left=Side(style='thin', color='D9E2F3'),
        right=Side(style='thin', color='D9E2F3'),
        top=Side(style='thin', color='D9E2F3'),
        bottom=Side(style='thin', color='D9E2F3')
    )

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for index, row in enumerate(data, start=1):
        ws.append([
            index,
            f"DN-{row[0]:04d}",
            row[1],
            row[2],
            row[3],
            row[4],
            row[5],
            row[6],
            row[7] if row[7] else '-',
            row[8],
            row[9],
            row[10]
        ])

    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center')

    for cell in ws['K']:
        if cell.row >= 4:
            cell.number_format = '"Rp" #,##0'

    column_widths = {
        'A': 6,
        'B': 14,
        'C': 24,
        'D': 18,
        'E': 14,
        'F': 30,
        'G': 16,
        'H': 16,
        'I': 18,
        'J': 16,
        'K': 16,
        'L': 18
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='laporan_denda.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/export_rekap_koleksi_excel')
def export_rekap_koleksi_excel():
    if 'login' not in session:
        return redirect('/')

    if session['role'] != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect('/dashboard_user')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            COALESCE(kategori, 'Tanpa Kategori') AS kategori,
            COUNT(id_buku) AS total_judul,
            COALESCE(SUM(jumlah_eksemplar), 0) AS total_eksemplar,
            COALESCE(SUM(stok_tersedia), 0) AS tersedia,
            COALESCE(SUM(jumlah_eksemplar - stok_tersedia), 0) AS dipinjam,
            COUNT(CASE WHEN status_buku ILIKE '%rusak%' THEN 1 END) AS rusak,
            COUNT(CASE WHEN status_buku ILIKE '%hilang%' THEN 1 END) AS hilang
        FROM buku
        GROUP BY kategori
        ORDER BY kategori ASC
    """)

    data = cursor.fetchall()

    cursor.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Koleksi"

    ws.merge_cells('A1:H1')
    ws['A1'] = 'REKAP KOLEKSI BUKU PERPUSTAKAAN'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = [
        'No',
        'Kategori',
        'Total Judul',
        'Total Eksemplar',
        'Tersedia',
        'Dipinjam',
        'Rusak',
        'Hilang'
    ]

    ws.append([])
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="7C3AED")
    header_font = Font(color="FFFFFF", bold=True)

    border = Border(
        left=Side(style='thin', color='D9E2F3'),
        right=Side(style='thin', color='D9E2F3'),
        top=Side(style='thin', color='D9E2F3'),
        bottom=Side(style='thin', color='D9E2F3')
    )

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    total_judul = 0
    total_eksemplar = 0
    total_tersedia = 0
    total_dipinjam = 0
    total_rusak = 0
    total_hilang = 0

    for index, row in enumerate(data, start=1):
        ws.append([
            index,
            row[0],
            row[1],
            row[2],
            row[3],
            row[4],
            row[5],
            row[6]
        ])

        total_judul += row[1]
        total_eksemplar += row[2]
        total_tersedia += row[3]
        total_dipinjam += row[4]
        total_rusak += row[5]
        total_hilang += row[6]

    ws.append([
        '',
        'TOTAL',
        total_judul,
        total_eksemplar,
        total_tersedia,
        total_dipinjam,
        total_rusak,
        total_hilang
    ])

    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center')

    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EDE9FE")

    column_widths = {
        'A': 6,
        'B': 24,
        'C': 16,
        'D': 18,
        'E': 14,
        'F': 14,
        'G': 12,
        'H': 12
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='rekap_koleksi_buku.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
@app.route('/proses_pengembalian/<int:id>')
def proses_pengembalian(id):
    if 'login' not in session:
        return redirect('/')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT 
            p.id_buku,
            p.tanggal_jatuh_tempo,
            COALESCE(pp.denda_per_hari, 1000)
        FROM peminjaman p
        JOIN users u ON p.id_user = u.id_user
        LEFT JOIN pengaturan_peminjaman pp ON pp.role = u.role::text
        WHERE p.id_peminjaman = %s
    """, (id,))

    data = cursor.fetchone()

    if data:
        id_buku = data[0]
        denda_per_hari = data[2]

        cursor.execute("""
            UPDATE peminjaman
            SET 
                tanggal_kembali = CURRENT_DATE,
                status_peminjaman = 'dikembalikan',
                denda = GREATEST(CURRENT_DATE - tanggal_jatuh_tempo, 0) * %s,
                updated_at = NOW()
            WHERE id_peminjaman = %s
        """, (denda_per_hari, id))

        cursor.execute("""
            UPDATE buku
            SET stok_tersedia = stok_tersedia + 1
            WHERE id_buku = %s
        """, (id_buku,))

        conn.commit()
        flash('Buku berhasil dikembalikan', 'success')

    cursor.close()
    conn.close()

    return redirect('/kelola_pengembalian')

@app.route('/export_laporan_peminjaman_excel')
def export_laporan_peminjaman_excel():

    if 'login' not in session:
        return redirect('/')

    if session['role'] != 'admin':
        flash("Akses ditolak", "danger")
        return redirect('/dashboard_user')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            p.id_peminjaman,
            u.nama,
            u.nomor_induk,
            b.judul_buku,
            b.penulis,
            p.tanggal_pinjam,
            p.tanggal_jatuh_tempo,
            p.tanggal_kembali,

            CASE
                WHEN p.tanggal_kembali IS NULL
                     AND CURRENT_DATE > p.tanggal_jatuh_tempo
                    THEN 'Terlambat'
                WHEN p.status_peminjaman='dikembalikan'
                    THEN 'Dikembalikan'
                ELSE 'Dipinjam'
            END AS status,

            CASE
                WHEN p.tanggal_kembali IS NULL
                     AND CURRENT_DATE > p.tanggal_jatuh_tempo
                    THEN CURRENT_DATE-p.tanggal_jatuh_tempo
                WHEN p.tanggal_kembali IS NOT NULL
                     AND p.tanggal_kembali>p.tanggal_jatuh_tempo
                    THEN p.tanggal_kembali-p.tanggal_jatuh_tempo
                ELSE 0
            END AS hari_terlambat,

            CASE
                WHEN p.tanggal_kembali IS NULL
                     AND CURRENT_DATE>p.tanggal_jatuh_tempo
                    THEN
                        (CURRENT_DATE-p.tanggal_jatuh_tempo)
                        *COALESCE(pp.denda_per_hari,1000)

                ELSE COALESCE(p.denda,0)
            END AS denda

        FROM peminjaman p
        JOIN users u
        ON u.id_user=p.id_user

        JOIN buku b
        ON b.id_buku=p.id_buku

        LEFT JOIN pengaturan_peminjaman pp
        ON pp.role=u.role::text

        ORDER BY p.id_peminjaman DESC
    """)

    data = cursor.fetchall()

    cursor.close()
    conn.close()

    wb = Workbook()

    ws = wb.active
    ws.title = "Rekap Peminjaman"

    ws.merge_cells("A1:K1")

    ws["A1"] = "LAPORAN REKAP PEMINJAMAN PERPUSTAKAAN"

    ws["A1"].font = Font(
        bold=True,
        size=14
    )

    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])

    headers = [
        "No",
        "ID",
        "Nama Anggota",
        "Nomor Induk",
        "Judul Buku",
        "Penulis",
        "Tanggal Pinjam",
        "Jatuh Tempo",
        "Tanggal Kembali",
        "Status",
        "Hari Terlambat",
        "Denda"
    ]

    ws.append(headers)

    fill = PatternFill(
        "solid",
        fgColor="2563EB"
    )

    font = Font(
        bold=True,
        color="FFFFFF"
    )

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for c in ws[3]:
        c.fill = fill
        c.font = font
        c.border = border
        c.alignment = Alignment(horizontal="center")

    total_denda = 0

    for i,row in enumerate(data,1):

        total_denda += row[10]

        ws.append([
            i,
            f"PJM-{row[0]:05d}",
            row[1],
            row[2],
            row[3],
            row[4],
            row[5],
            row[6],
            row[7] if row[7] else "-",
            row[8],
            row[9],
            row[10]
        ])

    ws.append([])

    ws.append([
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "TOTAL DENDA",
        "",
        total_denda
    ])

    for row in ws.iter_rows(min_row=4):

        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for cell in ws["L"]:
        if cell.row >= 4:
            cell.number_format = '"Rp" #,##0'

    width = {
        "A":6,
        "B":12,
        "C":25,
        "D":18,
        "E":35,
        "F":22,
        "G":18,
        "H":18,
        "I":18,
        "J":15,
        "K":15,
        "L":18
    }

    for k,v in width.items():
        ws.column_dimensions[k].width = v

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="laporan_rekap_peminjaman.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/export_laporan_buku_excel')
def export_laporan_buku_excel():
    if 'login' not in session:
        return redirect('/')

    if session['role'] != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect('/dashboard_user')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            b.id_buku,
            b.judul_buku,
            b.penulis,
            b.penerbit,
            b.tahun_terbit,
            b.isbn,
            b.bahasa,
            b.kategori,
            b.genre,
            COALESCE(b.jumlah_eksemplar, 0) AS jumlah_eksemplar,
            COALESCE(b.stok_tersedia, 0) AS stok_tersedia,
            COUNT(p.id_peminjaman) AS jumlah_dipinjam,
            b.status_buku,
            b.rak_lokasi,
            b.created_at
        FROM buku b
        LEFT JOIN peminjaman p ON b.id_buku = p.id_buku
        GROUP BY b.id_buku
        ORDER BY b.id_buku ASC
    """)

    data_buku = cursor.fetchall()

    cursor.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Buku"

    ws.merge_cells('A1:O1')
    ws['A1'] = 'LAPORAN DATA BUKU PERPUSTAKAAN'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = [
        'No', 'ID Buku', 'Judul Buku', 'Penulis', 'Penerbit',
        'Tahun Terbit', 'ISBN', 'Bahasa', 'Kategori', 'Genre',
        'Jumlah Eksemplar', 'Stok Tersedia', 'Jumlah Dipinjam',
        'Status Buku', 'Rak Lokasi'
    ]

    ws.append([])
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin', color='D9E2F3'),
        right=Side(style='thin', color='D9E2F3'),
        top=Side(style='thin', color='D9E2F3'),
        bottom=Side(style='thin', color='D9E2F3')
    )

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for index, row in enumerate(data_buku, start=1):
        ws.append([
            index,
            row[0],
            row[1],
            row[2],
            row[3],
            row[4],
            row[5],
            row[6],
            row[7],
            row[8],
            row[9],
            row[10],
            row[11],
            row[12],
            row[13]
        ])

    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center')

    column_widths = {
        'A': 6, 'B': 10, 'C': 30, 'D': 24, 'E': 24,
        'F': 14, 'G': 20, 'H': 14, 'I': 20, 'J': 20,
        'K': 18, 'L': 16, 'M': 18, 'N': 18, 'O': 14
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='laporan_buku.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def ambil_kata_kunci(query):
    query = query.lower().replace('?', '').replace('!', '').replace(',', '')

    if 'tema ' in query:
        return query.split('tema ', 1)[1].split()[0]

    if 'tentang ' in query:
        return query.split('tentang ', 1)[1].split()[0]

    return query.strip()

@app.route('/proses_rekomendasi_ai', methods=['POST'])
def proses_rekomendasi_ai():
    if 'login' not in session:
        return redirect('/')

    query_asli = request.form.get('query', '').strip().lower()

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM buku")
    semua_buku = cursor.fetchall()

    hasil = []

    for b in semua_buku:
        teks_buku = f"""
        {b[1]} {b[2]} {b[7]} {b[8]} {b[9]} {b[15]}
        """.lower()

        score = 0
        query_words = query_asli.split()

        for w in query_words:
            if w in teks_buku:
                score += 1

        if score > 0:
            hasil.append((score, b))

    hasil.sort(key=lambda x: x[0], reverse=True)

    rekomendasi = [x[1] for x in hasil[:5]]

    cursor.close()
    conn.close()

    return render_template(
        'member/rekomendasi_ai.html',
        nama=session['nama'],
        role=session['role'],
        query=query_asli,
        rekomendasi=rekomendasi
    )

@app.route('/pengaturan')
def pengaturan():
    if 'login' not in session:
        return redirect('/')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT 
            id_user, nama, email, role, foto_profil,
            nomor_induk, kelas, mata_pelajaran
        FROM users
        WHERE id_user = %s
    """, (session['id_user'],))

    user = cursor.fetchone()

    cursor.close()
    conn.close()

    if user is None:
        flash('Data pengguna tidak ditemukan', 'danger')
        return redirect('/dashboard_user')

    return render_template(
        'member/pengaturan.html',
        nama=session['nama'],
        user=user
    )

@app.route('/kelola_anggota')
def kelola_anggota():
    if 'login' not in session:
        return redirect('/')

    if session['role'] != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect('/dashboard_user')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM users")
    total_anggota = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM users WHERE aktif = true")
    anggota_aktif = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM users WHERE aktif = false")
    anggota_nonaktif = cursor.fetchone()[0]

    cursor.execute("""
        SELECT id_user, nama, email, password, role, aktif, foto_profil
        FROM users
        ORDER BY id_user DESC
    """)
    
    users = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        'admin/kelola_anggota.html',
        nama=session['nama'],
        users=users,
        total_anggota=total_anggota,
        anggota_aktif=anggota_aktif,
        anggota_nonaktif=anggota_nonaktif
    )

@app.route('/edit_anggota/<int:id>')
def edit_anggota(id):
    if 'login' not in session:
        return redirect('/')

    if session['role'] != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect('/dashboard_user')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT 
            id_user, nama, email, password, role, foto_profil, aktif,
            nomor_induk, kelas, mata_pelajaran, jenis_kelamin,
            tanggal_lahir, no_hp, alamat
        FROM users
        WHERE id_user = %s
    """, (id,))

    user = cursor.fetchone()

    cursor.close()
    conn.close()

    if user is None:
        flash('Data anggota tidak ditemukan', 'danger')
        return redirect('/kelola_anggota')

    return render_template(
        'admin/tambah_anggota.html',
        nama=session['nama'],
        mode='edit',
        user=user
    )

@app.route('/update_anggota/<int:id>', methods=['POST'])
def update_anggota(id):
    if 'login' not in session:
        return redirect('/')

    if session['role'] != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect('/dashboard_user')

    nama = request.form['nama']
    email = request.form['email']
    password = request.form.get('password')
    role = request.form['role']
    nomor_induk = request.form['nomor_induk']
    kelas = request.form.get('kelas') or None
    mata_pelajaran = request.form.get('mata_pelajaran') or None
    jenis_kelamin = request.form['jenis_kelamin']
    tanggal_lahir = request.form.get('tanggal_lahir') or None
    no_hp = request.form.get('no_hp')
    alamat = request.form.get('alamat')
    aktif = True if request.form['aktif'] == 'true' else False

    file = request.files.get('foto_profil')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT foto_profil, password
        FROM users
        WHERE id_user = %s
    """, (id,))

    old_user = cursor.fetchone()

    if old_user is None:
        cursor.close()
        conn.close()
        flash('Data anggota tidak ditemukan', 'danger')
        return redirect('/kelola_anggota')

    nama_file = old_user[0]
    old_password = old_user[1]

    if file and file.filename != "":
        nama_file = upload_sampul_ke_supabase(file)

    final_password = password if password else old_password

    cursor.execute("""
        UPDATE users
        SET
            nama = %s,
            email = %s,
            password = %s,
            role = %s,
            foto_profil = %s,
            aktif = %s,
            nomor_induk = %s,
            kelas = %s,
            mata_pelajaran = %s,
            jenis_kelamin = %s,
            tanggal_lahir = %s,
            no_hp = %s,
            alamat = %s,
            updated_at = NOW()
        WHERE id_user = %s
    """, (
        nama, email, final_password, role, nama_file, aktif,
        nomor_induk, kelas, mata_pelajaran, jenis_kelamin,
        tanggal_lahir, no_hp, alamat, id
    ))

    conn.commit()
    cursor.close()
    conn.close()

    flash('Data anggota berhasil diperbarui', 'success')
    return redirect('/kelola_anggota')
        
@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')

@app.route('/tambah_anggota')
def tambah_anggota():
    if 'login' not in session:
        return redirect('/')

    if session['role'] != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect('/dashboard_user')

    return render_template(
        'admin/tambah_anggota.html',
        nama=session['nama']
    )

@app.route('/simpan_anggota', methods=['POST'])
def simpan_anggota():
    if 'login' not in session:
        return redirect('/')

    if session['role'] != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect('/dashboard_user')

    nama = request.form['nama']
    email = request.form['email']
    password = request.form['password'] if request.form['password'] else '123456'
    role = request.form['role']
    nomor_induk = request.form['nomor_induk']
    kelas = request.form.get('kelas') or None
    mata_pelajaran = request.form.get('mata_pelajaran') or None
    jenis_kelamin = request.form['jenis_kelamin']
    tanggal_lahir = request.form.get('tanggal_lahir') or None
    no_hp = request.form.get('no_hp')
    alamat = request.form.get('alamat')
    aktif = True if request.form['aktif'] == 'true' else False

    file = request.files.get('foto_profil')
    nama_file = None

    if file and file.filename != "":
        nama_file = secure_filename(file.filename)
        file_path = os.path.join(PROFILE_UPLOAD_FOLDER, nama_file)
        file.save(file_path)

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO users (
            nama, email, password, role, foto_profil, aktif,
            nomor_induk, kelas, mata_pelajaran, jenis_kelamin,
            tanggal_lahir, no_hp, alamat
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        nama, email, password, role, nama_file, aktif,
        nomor_induk, kelas, mata_pelajaran, jenis_kelamin,
        tanggal_lahir, no_hp, alamat
    ))

    conn.commit()
    cursor.close()
    conn.close()

    flash('Anggota berhasil ditambahkan', 'success')
    return redirect('/kelola_anggota')

@app.route('/tambah_peminjaman')
def tambah_peminjaman():
    if 'login' not in session:
        return redirect('/')

    if session['role'] != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect('/dashboard_user')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id_user, nama, email, nomor_induk, kelas, aktif
        FROM users
        WHERE role = 'siswa' AND aktif = true
        ORDER BY nama ASC
    """)
    anggota = cursor.fetchall()

    cursor.execute("""
        SELECT id_buku, judul_buku, penulis, kategori, stok_tersedia, sampul
        FROM buku
        WHERE stok_tersedia > 0
        ORDER BY judul_buku ASC
    """)
    buku = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        'admin/tambah_peminjaman.html',
        mode='tambah',
        peminjaman=None,
        nama=session['nama'],
        anggota=anggota,
        buku=buku
    )

@app.route('/simpan_peminjaman', methods=['POST'])
def simpan_peminjaman():
    if 'login' not in session:
        return redirect('/')

    if session['role'] != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect('/dashboard_user')

    id_user = request.form['id_user']
    id_buku = request.form['id_buku']
    tanggal_pinjam = request.form['tanggal_pinjam']
    tanggal_jatuh_tempo = request.form['tanggal_jatuh_tempo']
    durasi_hari = request.form['durasi_hari']
    catatan = request.form.get('catatan')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO peminjaman (
            id_user, id_buku, tanggal_pinjam, tanggal_jatuh_tempo,
            durasi_hari, catatan, status_peminjaman, status_draft
        )
        VALUES (%s, %s, %s, %s, %s, %s, 'dipinjam', false)
    """, (
        id_user, id_buku, tanggal_pinjam, tanggal_jatuh_tempo,
        durasi_hari, catatan
    ))

    cursor.execute("""
        UPDATE buku
        SET stok_tersedia = stok_tersedia - 1
        WHERE id_buku = %s AND stok_tersedia > 0
    """, (id_buku,))

    conn.commit()
    cursor.close()
    conn.close()

    flash('Peminjaman berhasil disimpan', 'success')
    return redirect('/kelola_peminjaman')

@app.route('/kelola_peminjaman')
def kelola_peminjaman():
    if 'login' not in session:
        return redirect('/')

    if session['role'] != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect('/dashboard_user')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM peminjaman")
    total_transaksi = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM peminjaman
        WHERE status_peminjaman = 'dipinjam'
        AND tanggal_kembali IS NULL
        AND CURRENT_DATE <= tanggal_jatuh_tempo
    """)
    total_dipinjam = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM peminjaman
        WHERE status_peminjaman = 'dikembalikan'
    """)
    total_dikembalikan = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM peminjaman
        WHERE tanggal_kembali IS NULL
        AND CURRENT_DATE > tanggal_jatuh_tempo
    """)
    total_terlambat = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM peminjaman
        WHERE tanggal_pinjam = CURRENT_DATE
    """)
    total_hari_ini = cursor.fetchone()[0]

    cursor.execute("""
        SELECT 
            p.id_peminjaman,
            u.nama,
            u.nomor_induk,
            b.judul_buku,
            b.penulis,
            p.tanggal_pinjam,
            p.tanggal_jatuh_tempo,

            CASE
                WHEN p.tanggal_kembali IS NULL
                     AND CURRENT_DATE > p.tanggal_jatuh_tempo
                    THEN 'terlambat'
                ELSE p.status_peminjaman
            END AS status_peminjaman,

            CASE
                WHEN p.tanggal_kembali IS NULL
                     AND CURRENT_DATE > p.tanggal_jatuh_tempo
                    THEN (CURRENT_DATE - p.tanggal_jatuh_tempo) * COALESCE(pp.denda_per_hari, 1000)
                ELSE COALESCE(p.denda, 0)
            END AS denda,

            p.tanggal_kembali,

            GREATEST(CURRENT_DATE - p.tanggal_jatuh_tempo, 0) AS hari_terlambat

        FROM peminjaman p
        JOIN users u ON p.id_user = u.id_user
        JOIN buku b ON p.id_buku = b.id_buku
        LEFT JOIN pengaturan_peminjaman pp ON pp.role = u.role::text
        ORDER BY p.id_peminjaman DESC
    """)
    peminjaman = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        'admin/kelola_peminjaman.html',
        nama=session['nama'],
        peminjaman=peminjaman,
        total_transaksi=total_transaksi,
        total_dipinjam=total_dipinjam,
        total_dikembalikan=total_dikembalikan,
        total_terlambat=total_terlambat,
        total_hari_ini=total_hari_ini
    )

@app.route('/edit_peminjaman/<int:id>')
def edit_peminjaman(id):
    if 'login' not in session:
        return redirect('/')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id_peminjaman, id_user, id_buku, tanggal_pinjam,
               tanggal_jatuh_tempo, durasi_hari, catatan
        FROM peminjaman
        WHERE id_peminjaman = %s
    """, (id,))
    peminjaman = cursor.fetchone()

    cursor.execute("""
        SELECT id_user, nama, email, nomor_induk, kelas, aktif
        FROM users
        WHERE role = 'siswa' AND aktif = true
        ORDER BY nama ASC
    """)
    anggota = cursor.fetchall()

    cursor.execute("""
        SELECT id_buku, judul_buku, penulis, kategori, stok_tersedia, sampul
        FROM buku
        ORDER BY judul_buku ASC
    """)
    buku = cursor.fetchall()

    cursor.close()
    conn.close()

    if peminjaman is None:
        flash('Data peminjaman tidak ditemukan', 'danger')
        return redirect('/kelola_peminjaman')

    return render_template(
        'admin/tambah_peminjaman.html',
        mode='edit',
        peminjaman=peminjaman,
        anggota=anggota,
        buku=buku,
        nama=session['nama']
    )

@app.route('/update_peminjaman/<int:id>', methods=['POST'])
def update_peminjaman(id):
    if 'login' not in session:
        return redirect('/')

    id_user = request.form['id_user']
    id_buku = request.form['id_buku']
    tanggal_pinjam = request.form['tanggal_pinjam']
    tanggal_jatuh_tempo = request.form['tanggal_jatuh_tempo']
    durasi_hari = request.form['durasi_hari']
    catatan = request.form.get('catatan')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE peminjaman
        SET id_user = %s,
            id_buku = %s,
            tanggal_pinjam = %s,
            tanggal_jatuh_tempo = %s,
            durasi_hari = %s,
            catatan = %s,
            updated_at = NOW()
        WHERE id_peminjaman = %s
    """, (
        id_user, id_buku, tanggal_pinjam,
        tanggal_jatuh_tempo, durasi_hari,
        catatan, id
    ))

    conn.commit()
    cursor.close()
    conn.close()

    flash('Data peminjaman berhasil diperbarui', 'success')
    return redirect('/kelola_peminjaman')

@app.route('/pengaturan_admin')
def pengaturan_admin():
    if 'login' not in session:
        return redirect('/')

    if session['role'] != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect('/dashboard_user')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id_user, nama, email, role, foto_profil
        FROM users
        WHERE id_user = %s
    """, (session['id_user'],))
    admin = cursor.fetchone()

    cursor.execute("""
        SELECT id_pengaturan, role, maksimal_peminjaman, durasi_hari, denda_per_hari
        FROM pengaturan_peminjaman
        WHERE role = 'siswa'
        LIMIT 1
    """)
    pengaturan_pinjam = cursor.fetchone()

    cursor.execute("""
        SELECT id_pengaturan, nama_perpustakaan, alamat, no_telepon, email
        FROM pengaturan_perpustakaan
        LIMIT 1
    """)
    pengaturan_perpus = cursor.fetchone()

    cursor.close()
    conn.close()

    return render_template(
        'admin/pengaturan_admin.html',
        nama=session['nama'],
        admin=admin,
        pengaturan_pinjam=pengaturan_pinjam,
        pengaturan_perpus=pengaturan_perpus
    )
    
@app.route('/update_profil_admin', methods=['POST'])
def update_profil_admin():
    if 'login' not in session:
        return redirect('/')

    try:
        nama = request.form['nama']
        email = request.form['email']
        file = request.files.get('foto_profil')

        conn = get_db_connection()
        cursor = conn.cursor()

        # Folder upload foto profil
        upload_folder = os.path.join(BASE_DIR, 'static', 'images')
        os.makedirs(upload_folder, exist_ok=True)

        if file and file.filename:
            nama_file = secure_filename(file.filename)

            file_path = os.path.join(upload_folder, nama_file)

            # Simpan file
            file.save(file_path)

            # Update database termasuk foto
            cursor.execute("""
                UPDATE users
                SET nama = %s,
                    email = %s,
                    foto_profil = %s,
                    updated_at = NOW()
                WHERE id_user = %s
            """, (
                nama,
                email,
                nama_file,
                session['id_user']
            ))

        else:
            # Update tanpa foto
            cursor.execute("""
                UPDATE users
                SET nama = %s,
                    email = %s,
                    updated_at = NOW()
                WHERE id_user = %s
            """, (
                nama,
                email,
                session['id_user']
            ))

        conn.commit()

        cursor.close()
        conn.close()

        session['nama'] = nama
        session['email'] = email

        flash('Profil admin berhasil diperbarui', 'success')

    except Exception as e:
        print("ERROR UPDATE PROFIL:", e)
        flash(f'Gagal memperbarui profil: {str(e)}', 'danger')

    return redirect('/pengaturan_admin')

@app.route('/update_pengaturan_perpustakaan', methods=['POST'])
def update_pengaturan_perpustakaan():
    if 'login' not in session:
        return redirect('/')

    nama_perpustakaan = request.form['nama_perpustakaan']
    alamat = request.form['alamat']
    no_telepon = request.form['no_telepon']
    email = request.form['email']

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT id_pengaturan FROM pengaturan_perpustakaan LIMIT 1")
    data = cursor.fetchone()

    if data:
        cursor.execute("""
            UPDATE pengaturan_perpustakaan
            SET nama_perpustakaan = %s,
                alamat = %s,
                no_telepon = %s,
                email = %s
            WHERE id_pengaturan = %s
        """, (nama_perpustakaan, alamat, no_telepon, email, data[0]))
    else:
        cursor.execute("""
            INSERT INTO pengaturan_perpustakaan
            (nama_perpustakaan, alamat, no_telepon, email)
            VALUES (%s, %s, %s, %s)
        """, (nama_perpustakaan, alamat, no_telepon, email))

    conn.commit()
    cursor.close()
    conn.close()

    flash('Pengaturan perpustakaan berhasil diperbarui', 'success')
    return redirect('/pengaturan_admin')

@app.route('/update_profil_user', methods=['POST'])
def update_profil_user():
    if 'login' not in session:
        return redirect('/')

    try:
        nama = request.form['nama']
        email = request.form['email']
        nomor_induk = request.form.get('nomor_induk')
        kelas = request.form.get('kelas') or None
        mata_pelajaran = request.form.get('mata_pelajaran') or None
        file = request.files.get('foto_profil')

        conn = get_db_connection()
        cursor = conn.cursor()

        upload_folder = os.path.join(BASE_DIR, 'static', 'images')
        os.makedirs(upload_folder, exist_ok=True)

        if file and file.filename != "":
            nama_file = secure_filename(file.filename)
            file_path = os.path.join(upload_folder, nama_file)
            file.save(file_path)

            cursor.execute("""
                UPDATE users
                SET nama = %s,
                    email = %s,
                    nomor_induk = %s,
                    kelas = %s,
                    mata_pelajaran = %s,
                    foto_profil = %s,
                    updated_at = NOW()
                WHERE id_user = %s
            """, (
                nama, email, nomor_induk, kelas,
                mata_pelajaran, nama_file, session['id_user']
            ))
        else:
            cursor.execute("""
                UPDATE users
                SET nama = %s,
                    email = %s,
                    nomor_induk = %s,
                    kelas = %s,
                    mata_pelajaran = %s,
                    updated_at = NOW()
                WHERE id_user = %s
            """, (
                nama, email, nomor_induk, kelas,
                mata_pelajaran, session['id_user']
            ))

        conn.commit()
        cursor.close()
        conn.close()

        session['nama'] = nama
        session['email'] = email

        flash('Profil berhasil diperbarui', 'success')

    except Exception as e:
        print("ERROR UPDATE PROFIL USER:", e)
        flash(f'Gagal memperbarui profil: {str(e)}', 'danger')

    return redirect('/pengaturan')

@app.route('/update_password_user', methods=['POST'])
def update_password_user():
    if 'login' not in session:
        return redirect('/')

    password_lama = request.form['password_lama']
    password_baru = request.form['password_baru']
    konfirmasi_password = request.form['konfirmasi_password']

    if password_baru != konfirmasi_password:
        flash('Konfirmasi password tidak sama', 'danger')
        return redirect('/pengaturan')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT password
        FROM users
        WHERE id_user = %s
    """, (session['id_user'],))

    user = cursor.fetchone()

    if user is None or user[0] != password_lama:
        cursor.close()
        conn.close()
        flash('Password lama salah', 'danger')
        return redirect('/pengaturan')

    cursor.execute("""
        UPDATE users
        SET password = %s,
            updated_at = NOW()
        WHERE id_user = %s
    """, (password_baru, session['id_user']))

    conn.commit()
    cursor.close()
    conn.close()

    flash('Password berhasil diperbarui', 'success')
    return redirect('/pengaturan')

@app.route('/laporan')
def laporan():
    return render_template('admin/laporan.html')

@app.route('/rekomendasi_ai')
def rekomendasi_ai():
    if 'login' not in session:
        return redirect('/')

    return render_template(
        'member/rekomendasi_ai.html',
        nama=session['nama'],
        role=session['role'],
        mata_pelajaran=None,
        kelas=None,
        rekomendasi=[],
        query=None
    )

@app.route('/laporan_denda')
def laporan_denda():
    return render_template('admin/laporan_denda.html')

@app.route('/laporan_buku')
def laporan_buku():
    return render_template('admin/laporan_buku.html')

@app.route('/rekap_koleksi')
def rekap_koleksi():
    return render_template('admin/rekap_koleksi.html')
    
@app.route('/update_pengaturan_peminjaman', methods=['POST'])
def update_pengaturan_peminjaman():
    if 'login' not in session:
        return redirect('/')

    maksimal_peminjaman = request.form['maksimal_peminjaman']
    durasi_hari = request.form['durasi_hari']
    denda_per_hari = request.form['denda_per_hari']

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE pengaturan_peminjaman
        SET maksimal_peminjaman = %s,
            durasi_hari = %s,
            denda_per_hari = %s
        WHERE role = 'siswa'
    """, (maksimal_peminjaman, durasi_hari, denda_per_hari))

    conn.commit()
    cursor.close()
    conn.close()

    flash('Pengaturan peminjaman berhasil diperbarui', 'success')
    return redirect('/pengaturan_admin')


@app.route('/update_password_admin', methods=['POST'])
def update_password_admin():
    if 'login' not in session:
        return redirect('/')

    password_lama = request.form['password_lama']
    password_baru = request.form['password_baru']
    konfirmasi_password = request.form['konfirmasi_password']

    if password_baru != konfirmasi_password:
        flash('Konfirmasi password tidak sama', 'danger')
        return redirect('/pengaturan_admin')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT password FROM users
        WHERE id_user = %s
    """, (session['id_user'],))
    user = cursor.fetchone()

    if user is None or user[0] != password_lama:
        cursor.close()
        conn.close()
        flash('Password lama salah', 'danger')
        return redirect('/pengaturan_admin')

    cursor.execute("""
        UPDATE users
        SET password = %s
        WHERE id_user = %s
    """, (password_baru, session['id_user']))

    conn.commit()
    cursor.close()
    conn.close()

    flash('Password berhasil diperbarui', 'success')
    return redirect('/pengaturan_admin')
    
if __name__ == '__main__':
    app.run(debug=True, use_reloader=False)